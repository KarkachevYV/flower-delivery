# orders/tasks.py 
print("Модуль orders.tasks загружен")

from fpdf import FPDF
from celery import shared_task
from django.core.management import call_command
from django.utils.timezone import now
from .models import Order, OrderItem
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import openpyxl
from django.db.models import Sum, Count, F
import os

def calculate_totals(completed_orders):
    """Вычисляет итоговые значения."""
    total_orders = completed_orders.count()
    total_items = OrderItem.objects.filter(order__in=completed_orders).count()
    total_revenue = completed_orders.aggregate(total=Sum(F('orderitem__quantity') * F('orderitem__price')))['total'] or 0
    return total_orders, total_items, total_revenue

@shared_task
def generate_daily_report_task():
    """Основная задача для генерации ежедневного отчёта."""
    today = now().date()
    print(f"Запуск задачи для отчёта за {today}")
    try:
        call_command('generate_daily_report')
        generate_pdf_report(today) 
        generate_excel_report(today) 
    except Exception as e:
        print(f"Ошибка при генерации отчётов: {e}")

# @shared_task
# def generate_pdf_report(today):
#     """Создаёт PDF-отчёт."""
#     try:
#         completed_orders = Order.objects.filter(status='completed', created_at__date=today)
#         total_orders, total_items, total_revenue = calculate_totals(completed_orders)

#         file_path = os.path.join('reports', f"daily_report_{today}.pdf")
#         c = canvas.Canvas(file_path, pagesize=A4)

#         c.drawString(100, 800, f"Отчёт за {today}")
#         c.drawString(100, 780, f"Завершённых заказов: {total_orders}")
#         c.drawString(100, 760, f"Проданных товаров: {total_items}")
#         c.drawString(100, 740, f"Общая выручка: {total_revenue} руб.")

#         c.save()
#         print(f"PDF-отчёт сохранён: {file_path}")
#     except Exception as e:
#         print(f"Ошибка при создании PDF-отчёта: {e}")

# print("Функция generate_pdf_report объявлена")


@shared_task
def generate_pdf_report():
    try:
        today = now().date()
        completed_orders = Order.objects.filter(created_at__date=today)

        total_orders = completed_orders.count()
        total_items = OrderItem.objects.filter(order__in=completed_orders).count()
        total_revenue = OrderItem.objects.filter(order__in=completed_orders).aggregate(
            total=Sum(F('quantity') * F('price'))
        )['total'] or 0

        # Создание PDF
        pdf = FPDF()
        pdf.add_page()

        # Добавление шрифта с поддержкой кириллицы
        font_path = 'fonts/DejaVuSans.ttf'
        pdf.add_font('DejaVu', '', font_path, uni=True)
        pdf.set_font('DejaVu', '', 14)
        # pdf.add_font('FreeSerif', '', 'FreeSerif.ttf', uni=True)
        # pdf.set_font('FreeSerif', '', 12)


        pdf.cell(200, 10, txt="Отчёт по заказам за сегодня", ln=True, align='C')
        pdf.ln(10)
        pdf.cell(200, 10, txt=f"Заказов: {total_orders}", ln=True)
        pdf.cell(200, 10, txt=f"Товаров в заказах: {total_items}", ln=True)
        pdf.cell(200, 10, txt=f"Общая сумма: {total_revenue}₽", ln=True)

        # Путь для сохранения отчёта
        reports_dir = "reports"
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, f"daily_report_{today}.pdf")

        pdf.output(file_path)
        print(f"PDF-отчёт сохранён: {file_path}")
        return file_path

    except Exception as e:
        print(f"Ошибка при создании PDF-отчёта: {e}")


@shared_task
def generate_excel_report(today):
    """Создаёт Excel-отчёт."""
    try:
        completed_orders = Order.objects.filter(status='completed', created_at__date=today)
        total_orders, total_items, total_revenue = calculate_totals(completed_orders)

        file_path = os.path.join('reports', f"daily_report_{today}.xlsx")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ежедневный отчёт"

        ws.append(["Дата", "Завершённые заказы", "Проданные товары", "Выручка"])
        ws.append([today, total_orders, total_items, total_revenue])

        wb.save(file_path)
        print(f"Excel-отчёт сохранён: {file_path}")
    except Exception as e:
        print(f"Ошибка при создании Excel-отчёта: {e}")

@shared_task
def generate_daily_report():
    """Генерирует текстовый отчёт и выводит его в консоль."""
    try:
        today = now().date()
        completed_orders = Order.objects.filter(status='completed', created_at__date=today)

        total_orders, _, total_revenue = calculate_totals(completed_orders)

        report = f"Отчёт за {today}:\nЗавершённых заказов: {total_orders}\nВыручка: {total_revenue}₽"
        print(report)
        return report
    except Exception as e:
        print(f"Ошибка при генерации текстового отчёта: {e}")